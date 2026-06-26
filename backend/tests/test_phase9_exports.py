import os, sys, os.path; sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
os.environ.update({"SUPABASE_DB_HOST":"x","SUPABASE_DB_USER":"x","SUPABASE_DB_PASSWORD":"x",
 "SECRET_KEY":"testsecret","ENCRYPTION_MASTER_KEY":"k","SUPABASE_URL":"https://x.supabase.co"})
import time
from app.services.whatif import WhatIfInputs, simulate
from app.services import export_service as ex
from app.core_tokens import make_download_token, verify_download_token

P=[];F=[]
def ck(n,c): (P if c else F).append(n); print(("PASS " if c else "FAIL ")+n)

# --- What-If math ---
# base 1,200,000; recover 50% => 600,000; over 3 months; cost 150,000 total
r = simulate(WhatIfInputs(base_amount_iqd=1_200_000, recovery_pct=50,
    implementation_months=3, implementation_cost_iqd=150_000, horizon_months=6))
ck("recovered = 600000", r.recovered_amount==600_000)
ck("monthly cost = 50000", r.monthly_implementation_cost==50_000)
# monthly cash flow during impl = 600000/3 - 50000 = 200000-50000 = 150000
ck("monthly cash flow = 150000", r.monthly_cash_flow_impact==150_000)
# net profit = 600000 - 150000 = 450000
ck("net profit = 450000", r.net_profit_impact==450_000)
ck("projection has 6 months", len(r.projection)==6)
# cumulative after 3 months = 3*150000 = 450000; stays flat after
ck("cumulative m3 = 450000", r.projection[2]["cumulative_cash_flow"]==450_000)
ck("cumulative m6 = 450000 (flat post-impl)", r.projection[5]["cumulative_cash_flow"]==450_000)
# clamp recovery
r2 = simulate(WhatIfInputs(base_amount_iqd=1000, recovery_pct=150, implementation_months=1, implementation_cost_iqd=0))
ck("recovery clamped to 100%", r2.recovered_amount==1000)

# --- download token ---
tok = make_download_token("exports/x.pdf", "comp1", ttl_seconds=900)
p = verify_download_token(tok)
ck("token verifies", p and p["path"]=="exports/x.pdf" and p["company_id"]=="comp1")
ck("tampered token rejected", verify_download_token(tok[:-3]+"abc") is None)
expired = make_download_token("exports/x.pdf", "c", ttl_seconds=-5)
ck("expired token rejected", verify_download_token(expired) is None)

# --- renderers produce valid files with Arabic ---
headers=["القسم","المبلغ (د.ع)","الوصف"]
rows=[["المشتريات",1500000,"دفع مكرر"],["المخازن",900000,"نقص جرد"]]
cert={"report_id":"r1","ledger_hash_at_generation":"abc123","digital_signature":"sig","algorithm":"HMAC-SHA256"}

xlsx=ex._render_excel("خريطة الهدر", headers, rows, cert)
ck("excel is xlsx (zip magic PK)", xlsx[:2]==b"PK" and len(xlsx)>2000)
pdf=ex._render_pdf("خريطة الهدر", headers, rows, cert)
ck("pdf magic %PDF", pdf[:4]==b"%PDF" and len(pdf)>1000)
png=ex._render_png("خريطة الهدر", headers, rows, cert)
ck("png magic", png[:8]==bytes([0x89,0x50,0x4e,0x47,0x0d,0x0a,0x1a,0x0a]) and len(png)>5000)

# verify excel RTL + cert sheet
import io
from openpyxl import load_workbook
wb=load_workbook(io.BytesIO(xlsx))
ck("excel sheet is RTL", wb["التقرير"].sheet_view.rightToLeft is True)
ck("excel has cert sheet", "شهادة عدم التلاعب" in wb.sheetnames)

print(f"\n=== {len(P)} passed, {len(F)} failed ===")
sys.exit(1 if F else 0)
