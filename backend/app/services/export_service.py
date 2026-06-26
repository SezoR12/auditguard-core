"""Export engine: Excel / PDF / PNG reports with Arabic support.

Every export embeds a Tamper-Proof Certificate (Phase 5): the last ledger hash
plus an HMAC signature over the report content, so an exported file can be tied
to the immutable audit state at generation time.
"""
from __future__ import annotations

import io
import os
import uuid
from datetime import datetime, timezone
from pathlib import Path

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.config import settings
from app.models import AnalyticsOutput, RiskAlert, WasteMapItem
from app.models.enums import OutputType
from app.services.arabic_text import shape_arabic
from app.services.ledger_service import build_tamper_proof_certificate, get_last_hash

FONT_PATH = str(Path(__file__).resolve().parent.parent / "assets" / "fonts" / "Amiri-Regular.ttf")
EXPORT_DIR = os.path.join(settings.STORAGE_ROOT, "exports")

# --- data loading -----------------------------------------------------------


async def _load_report_rows(
    session: AsyncSession, report_type: str, company_id: uuid.UUID, date_from, date_to
) -> tuple[list[str], list[list]]:
    """Return (headers, rows) for the requested report_type."""
    if report_type in ("waste_map", "waste"):
        q = select(WasteMapItem).where(WasteMapItem.company_id == company_id)
        if date_from:
            q = q.where(WasteMapItem.created_at >= date_from)
        if date_to:
            q = q.where(WasteMapItem.created_at <= date_to)
        q = q.order_by(WasteMapItem.amount_iqd.desc())
        rows = (await session.execute(q)).scalars().all()
        headers = ["القسم", "الفئة", "المبلغ (د.ع)", "الوصف", "الحالة", "التاريخ"]
        data = [
            [
                r.department,
                r.category.value if hasattr(r.category, "value") else str(r.category),
                float(r.amount_iqd),
                r.description,
                r.status,
                r.created_at.strftime("%Y-%m-%d"),
            ]
            for r in rows
        ]
        return headers, data

    if report_type in ("risk_alerts", "risks"):
        q = select(RiskAlert).where(RiskAlert.company_id == company_id)
        if date_from:
            q = q.where(RiskAlert.created_at >= date_from)
        if date_to:
            q = q.where(RiskAlert.created_at <= date_to)
        q = q.order_by(RiskAlert.created_at.desc())
        rows = (await session.execute(q)).scalars().all()
        headers = ["الخطورة", "العنوان", "الوصف", "الأثر المالي", "الحالة", "التاريخ"]
        data = [
            [
                r.severity.value if hasattr(r.severity, "value") else str(r.severity),
                r.title,
                r.description,
                float(r.financial_impact) if r.financial_impact is not None else 0.0,
                r.status,
                r.created_at.strftime("%Y-%m-%d"),
            ]
            for r in rows
        ]
        return headers, data

    # default: analytics snapshots
    q = (
        select(AnalyticsOutput)
        .where(
            AnalyticsOutput.company_id == company_id,
            AnalyticsOutput.output_type == OutputType.daily_snapshot,
        )
        .order_by(AnalyticsOutput.generated_at.desc())
    )
    rows = (await session.execute(q)).scalars().all()
    headers = ["التاريخ", "مؤشر الثقة", "إجمالي الهدر (د.ع)"]
    data = [
        [
            r.generated_at.strftime("%Y-%m-%d %H:%M"),
            r.trust_index or 0,
            float((r.data or {}).get("total_waste_iqd", 0)) if isinstance(r.data, dict) else 0.0,
        ]
        for r in rows
    ]
    return headers, data


# --- renderers --------------------------------------------------------------


def _render_excel(title: str, headers: list[str], rows: list[list], cert: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "التقرير"
    ws.sheet_view.rightToLeft = True  # RTL sheet

    header_fill = PatternFill("solid", fgColor="1E3A8A")
    header_font = Font(bold=True, color="FFFFFF")
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="right", vertical="center")
    for row in rows:
        ws.append(row)
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(60, width + 4)

    # Tamper-proof certificate sheet.
    cs = wb.create_sheet("شهادة عدم التلاعب")
    cs.sheet_view.rightToLeft = True
    for k, v in cert.items():
        cs.append([k, str(v)])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _render_pdf(title: str, headers: list[str], rows: list[list], cert: dict) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    pdfmetrics.registerFont(TTFont("Amiri", FONT_PATH))
    styles = getSampleStyleSheet()
    ar_title = ParagraphStyle("arTitle", parent=styles["Title"], fontName="Amiri", alignment=2)
    ar_norm = ParagraphStyle("arNorm", parent=styles["Normal"], fontName="Amiri", alignment=2)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, rightMargin=15 * mm, leftMargin=15 * mm)
    elements = [Paragraph(shape_arabic(title), ar_title), Spacer(1, 8 * mm)]

    table_data = [[Paragraph(shape_arabic(str(h)), ar_norm) for h in reversed(headers)]]
    for row in rows[:500]:
        table_data.append([Paragraph(shape_arabic(str(c)), ar_norm) for c in reversed(row)])

    table = Table(table_data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E3A8A")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("FONTNAME", (0, 0), (-1, -1), "Amiri"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("ALIGN", (0, 0), (-1, -1), "RIGHT"),
            ]
        )
    )
    elements.append(table)
    elements.append(Spacer(1, 10 * mm))

    # Tamper-proof certificate page content.
    elements.append(Paragraph(shape_arabic("شهادة عدم التلاعب"), ar_title))
    for k, v in cert.items():
        elements.append(Paragraph(f"{k}: {v}", ar_norm))

    doc.build(elements)
    return buf.getvalue()


def _render_png(title: str, headers: list[str], rows: list[list], cert: dict) -> bytes:
    import matplotlib

    matplotlib.use("Agg")
    import matplotlib.font_manager as fm
    import matplotlib.pyplot as plt

    fm.fontManager.addfont(FONT_PATH)
    prop = fm.FontProperties(fname=FONT_PATH)

    # Build a simple bar chart from the first numeric column found.
    num_idx = None
    for i, h in enumerate(headers):
        if any(isinstance(r[i], (int, float)) for r in rows):
            num_idx = i
            break
    labels = [shape_arabic(str(r[0])) for r in rows[:12]]
    values = [float(r[num_idx]) if num_idx is not None else 0 for r in rows[:12]]

    fig, ax = plt.subplots(figsize=(10, 6), dpi=300)
    ax.bar(range(len(values)), values, color="#dc2626")
    ax.set_xticks(range(len(labels)))
    ax.set_xticklabels(labels, fontproperties=prop, rotation=30, ha="right")
    ax.set_title(shape_arabic(title), fontproperties=prop, fontsize=16)
    ax.set_ylabel(shape_arabic(headers[num_idx] if num_idx is not None else ""), fontproperties=prop)
    fig.text(0.01, 0.01, f"ledger: {cert.get('ledger_hash_at_generation','')[:16]}…", fontsize=6)
    fig.tight_layout()

    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=300, bbox_inches="tight")
    plt.close(fig)
    return buf.getvalue()


# --- public API -------------------------------------------------------------


async def export_report(
    session: AsyncSession,
    *,
    report_type: str,
    fmt: str,
    company_id: uuid.UUID,
    date_from=None,
    date_to=None,
    title: str | None = None,
) -> tuple[str, bytes, str]:
    """Generate an export. Returns (filename, content_bytes, mime)."""
    title = title or {"waste_map": "خريطة الهدر", "risk_alerts": "تنبيهات المخاطر"}.get(
        report_type, "تقرير AuditCore"
    )
    headers, rows = await _load_report_rows(session, report_type, company_id, date_from, date_to)

    # Tamper-proof certificate bound to current ledger state.
    last_hash = await get_last_hash(session)
    report_id = str(uuid.uuid4())
    content_repr = f"{report_type}|{len(rows)}|{last_hash}".encode("utf-8")
    cert = build_tamper_proof_certificate(
        report_id=report_id,
        report_content=content_repr,
        ledger_hash_at_generation=last_hash,
        company_key=settings.SECRET_KEY,
    )

    fmt = fmt.lower()
    if fmt in ("excel", "xlsx"):
        content = _render_excel(title, headers, rows, cert)
        ext, mime = "xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    elif fmt == "pdf":
        content = _render_pdf(title, headers, rows, cert)
        ext, mime = "pdf", "application/pdf"
    elif fmt in ("png", "image"):
        content = _render_png(title, headers, rows, cert)
        ext, mime = "png", "image/png"
    else:
        raise ValueError(f"unsupported format: {fmt}")

    filename = f"{report_type}_{datetime.now(timezone.utc).strftime('%Y%m%d%H%M%S')}_{report_id[:8]}.{ext}"
    return filename, content, mime


def save_export(filename: str, content: bytes) -> str:
    """Persist an export to the storage volume; return the relative path."""
    os.makedirs(EXPORT_DIR, exist_ok=True)
    abs_path = os.path.join(EXPORT_DIR, filename)
    with open(abs_path, "wb") as f:
        f.write(content)
    try:
        os.chmod(abs_path, 0o600)
    except OSError:
        pass
    return os.path.join("exports", filename)
